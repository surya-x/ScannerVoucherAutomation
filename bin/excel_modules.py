from openpyxl import *
from bin.config import *
from datetime import datetime
import os


def write_excel(empID, list_of_data):
    filename = datetime.today().strftime('%Y%m%d') + ".xlsx"
    file_path = os.path.join(excel_path, filename)

    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    sheet = workbook.active

    row_count = sheet.max_row
    row_num = str(row_count+1)

    sheet["A" + row_num] = empID
    sheet["B" + row_num] = list_of_data[0]
    sheet["C" + row_num] = list_of_data[1]
    sheet["D" + row_num] = list_of_data[2]

    workbook.save(filename=file_path)
